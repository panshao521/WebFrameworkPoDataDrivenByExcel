# -*- coding: utf-8 -*-
"""
Time:     2021/7/21 21:39
Author:   panyuangao@foxmail.com
File:     Excel.py
Describe: 封装操作excel的方法
"""
from openpyxl import *
import os
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill
from openpyxl.styles import  colors,Font, Border,Side

class ExcelUtil():

    def __init__(self,excel_file_path):
        if os.path.exists(excel_file_path):
            self.excel_file_path = excel_file_path
            #加载excel文件
            self.wb = load_workbook(self.excel_file_path)
            #获取当前文件默认 sheet 的对象
            self.sheet = self.wb.active  #
        else:
            print("Excel 文件 %s 路径不存在" %excel_file_path)
            self.excel_file_path = None
            self.wb = None
            self.sheet = None

    # 获取excel所有的sheet名称
    def get_sheet_names(self):
        if self.wb is not None:
            return self.wb.sheetnames
        else:
            return None

    # 获取excel文件绝对路径
    def get_file_path(self):
        return self.excel_file_path

    # 通过sheet序号选择sheet
    def set_sheet_by_index(self,index):
        if not isinstance(index,int):
            print("您设定的sheet序号不是数字类型，请重新设定")
            return None
        elif not 0<=index<len(self.get_sheet_names()):
            print("您设定的sheet序号【%s】不存在，请重新设定！最大sheet编号是%s" %(index,len(self.get_sheet_names())-1))
            return None
        self.sheet = self.wb[self.get_sheet_names()[index]]
        return self.sheet

    # 通过sheet名称选择sheet
    def set_sheet_by_name(self, sheet_name):
        if sheet_name not in self.get_sheet_names():
            print("您设定的sheet名称【%s】不存在，请重新选择sheet名称" %sheet_name)
            return None
        self.sheet = self.wb[sheet_name]
        return self.sheet

    # 通过sheet名称创建一个sheet
    def create_sheet(self,sheet_name):
        if sheet_name in self.get_sheet_names():
            print("sheet名称【%s】已经存在，请重新设定sheet名称" %sheet_name)
            return
        self.wb.create_sheet(sheet_name)
        self.wb.save(self.get_file_path())

    # 获取sheet的最大行号
    def get_max_row_count(self):
        return self.sheet.max_row

    # 获取sheet的最大列号
    def get_max_col_count(self):
        return self.sheet.max_column

    #获得当前sheet中所有的单元格对象，返回一个带有多个子列表的列表
    def get_sheet_all_cells(self):
        cell_objects = [] #存储整个儿sheet的单元对象列表，每一行cell会存在一个子列表中
        for row in self.sheet.iter_rows():
            row_cell_objects = [] #存储某一行的所有cell对象
            for cell in row:
                row_cell_objects.append(cell)
            cell_objects.append(row_cell_objects)
        return cell_objects

    # 获取当前sheet中所有的单元格对象的值，返回一个带有多个子列表的列表
    def get_sheet_all_cell_values(self):
        values = []
        for row in self.get_sheet_all_cells():
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            values.append(row_values)
        return values

    #获取一个单元格的值
    def get_cell_value(self, row_no, col_no):
        if not (isinstance(row_no, int) or isinstance(col_no, int)):
            print("您输入的行号%s, 或者输入的列号%s不是整数" %(row_no, col_no))
            return None
        if not 0 <= row_no < self.get_max_row_count():
            print("您输入的行号%s不在合法的行号范围内【0 - %s】，请重新输入" %(row_no, self.get_max_row_count()-1))
            return None
        if not 0 <= col_no < self.get_max_col_count():
            print("您输入的列号%s不在合法的列号范围内【0 - %s】，请重新输入" %(col_no, self.get_max_col_count()-1))
            return None
        return self.sheet.cell(row_no+1,col_no+1).value

    # 获取某一行（从0开始）的单元格的值，放到列表中
    def get_row_values(self, row_no):
        if (not isinstance(row_no, int)):
            print("输入的行号%s不是整数" % (row_no))
            return None
        if not 0 <= row_no < self.get_max_row_count():
            print("您输入的行号%s不在合法的行号范围内【0 - %s】，请重新输入" %(row_no, self.get_max_row_count()-1))
            return None
        return self.get_sheet_all_cell_values()[row_no]

    # 获取某一列（从0开始）的单元格的值，放到列表中
    def get_col_values(self, col_no):
        if (not isinstance(col_no, int)):
            print("输入的行号%s不是整数" % (col_no))
            return None
        if not 0 <= col_no < self.get_max_col_count():
            print("您输入的列号%s不在合法的列号范围内【0 - %s】，请重新输入" %(col_no, self.get_max_col_count()-1))
            return None
        col_values = []
        for row in self.get_sheet_all_cell_values():
            for idx, cell in enumerate(row):
                if idx == col_no:
                    col_values.append(cell)
        return col_values

    #将多行数据写入到excel中，追加到excel 的最后面
    def write_lines_in_sheet(self,data,border_flag = True):
        if not isinstance(data,(list,tuple)):
            print("您写入的数据不是元组或列表类型，请重新设定")
            return
        for line in data: # data是需要包含多个子列表或子元组的类型
            if not isinstance(line, (list, tuple)):
                print("你写入的数据行,不是列表或者元组类型，请重新设定")
                return
            self.sheet.append(line)

        if border_flag:
            bd = Side(style='thin', color="000000") # 给excel设定黑色细边框
            for row in self.sheet.rows:
                for cell in row:
                    cell.border = Border(left=bd,top=bd,right=bd,bottom=bd)
        self.wb.save(self.get_file_path())

    #将一行内容追加到excel的最后面，设定是否有边框、字体的颜色（red、green，None--》黑色）、背景色
    def write_a_line_in_sheet(self, data, border_flag=True, font_color=None, fgcolor=None):
        '''
        :param data:数据
        :param border_flag:是否设置边框，True，False
        :param font_color:字体颜色
        :param fgcolor:填充背景颜色
        :return:
        '''
        if fgcolor is not None:
            fill = PatternFill(fill_type="solid", fgColor=fgcolor)
        else:
            fill = None
        if font_color is None:
            ft = None
        elif "red" in font_color:
            ft = Font(color=colors.RED)
        elif "green" in font_color:
            ft = Font(color=colors.GREEN)

        if not isinstance(data, (list, tuple)):
            print("你写入的数据行,不是列表或者元组类型，请重新设定")
            return
        first_line = self.sheet[1] # 获取第一行数据
        if len(first_line) == 1 and first_line[0].value is None: # 判断第一行数据是否为空
            rowNo = self.get_max_row_count()
        else:
            rowNo = self.get_max_row_count() + 1

        for idx, value in enumerate(data):
            # print(idx, value)
            if font_color is not None:
                if ("成功" in str(value)) or ("失败" in str(value)) and ft is not None: #包含“成功”、“失败”关键字才设定颜色
                    # 行号和列号都是从1开始
                    self.sheet.cell(row=rowNo, column=idx + 1).font = ft # 设定单元格字体颜色
            if fgcolor is not None:
                # 行号和列号都是从1开始
                self.sheet.cell(row=rowNo, column=idx + 1).fill = fill # 设定单元格填充的背景颜色
            self.sheet.cell(row=rowNo, column=idx + 1).value = value
        bd = Side(style='thin', color="000000") # 单元格边框样式，黑色细线
        if border_flag:
            for row in self.sheet.rows:
                for cell in row:
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd) # 设置单元格边框样式
        self.wb.save(self.get_file_path())

    # 写入某一列的值，列号从0开始
    def write_a_col_in_sheet(self, col_no, data, border_flag=True, font_color=None, fgcolor=None):
        if fgcolor is not None:
            fill = PatternFill(fill_type="solid", fgColor=fgcolor)
        else:
            fill = None
        if font_color is None:
            ft = None
        elif "red" in font_color:
            ft = Font(color=colors.RED)
        elif "green" in font_color:
            ft = Font(color=colors.GREEN)

        if not isinstance(data, (list, tuple)):
            print("你写入的数据列,不是列表或者元组类型，请重新设定")
            return

        bd = Side(style='thin', color="000000")
        for idx, value in enumerate(data):
            self.sheet.cell(row=idx + 1, column=col_no + 1).value = str(value)
            if font_color is not None:
                if ("成功" in str(value)) or ("失败" in str(value)) and ft is not None:
                    self.sheet.cell(row=idx + 1, column=col_no + 1).font = ft
            if fgcolor is not None:
                self.sheet.cell(row=idx + 1, column=col_no + 1).fill = fill
            if border_flag:
                self.sheet.cell(row=idx + 1, column=col_no + 1).border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.save()

    #写入某个单元格值，坐标从0行0列开始，字体可设置颜色红、绿、黑（默认是黑色），默认有边框
    def write_cell_value(self, row_no, col_no, value, color=None, border=True):
        if not (isinstance(row_no, int) or isinstance(col_no, int)):
            print("您输入的行号%s, 或者输入的列号%s不是整数" %(row_no, col_no))
            return None
        if color:
            if 'red' in color:
                self.sheet.cell(row=row_no+1, column=col_no+1).font = Font(color=colors.RED)
            elif 'green' in color:
                self.sheet.cell(row=row_no + 1, column=col_no + 1).font = Font(color=colors.GREEN)
            else:
                self.sheet.cell(row=row_no + 1, column=col_no + 1).font = Font(color=colors.BLACK)
        if border:
            bd = Side(style='thin',color='000000')
            self.sheet.cell(row=row_no + 1, column=col_no + 1).border = Border(left=bd,top=bd,right=bd,bottom=bd)
        self.sheet.cell(row=row_no + 1, column=col_no + 1).value = value

    def write_date_in_cell(self,row_no,col_num):
        from Util.DateStr import get_chinese_current_datetime
        date = get_chinese_current_datetime()
        self.write_cell_value(row_no,col_num,date)

    # 保存excel
    def save(self):
        self.wb.save(self.get_file_path())



if __name__ == "__main__":
    excel = ExcelUtil("/Users/panshao/Documents/test.xlsx")
    print(excel.get_sheet_names())
    print(excel.get_file_path())
    # print(excel.set_sheet_by_index(10))
    # print(excel.set_sheet_by_name("Sheet10"))
    # excel.create_sheet("测试数据")
    # print(excel.get_sheet_names())

    # excel.set_sheet_by_name("Sheet1")
    # print(excel.get_max_row_count())
    # print(excel.get_max_col_count())
    # print(excel.get_sheet_all_cells())
    # excel.write_lines_in_sheet([['你好', "不错", "厉害", "压力山大"], ['OK', "good"]])
    # print(excel.get_sheet_all_cell_values())

    # print(excel.set_sheet_by_name("Sheet1"))
    # excel.write_a_line_in_sheet(["你好", "wohao ", "ok", "成功"], font_color="red")
    excel.write_a_line_in_sheet(["你好", "wohao ", "ok", "失败"], font_color="green")
    excel.write_a_line_in_sheet(["你好", "wohao ", "ok", "成功"], font_color="red", fgcolor="CD9B9B")
    # print(excel.get_cell_value(4,3))
    # print(excel.get_row_value(5))
    # print(excel.get_col_value(4))
    # excel.write_cell_value(1,6,"我很开心",'green',False)
    # excel.write_a_col_in_sheet(2,[1,2,3,4,"成功",'red'],'True','red')
    excel.write_date_in_cell(4,8)


    excel.save()

